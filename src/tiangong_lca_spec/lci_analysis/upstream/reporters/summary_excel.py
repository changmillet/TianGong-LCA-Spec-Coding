"""Excel export for lifecycle flow prioritisation output."""

from __future__ import annotations

from itertools import chain
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from tiangong_lca_spec.lci_analysis.upstream.models import (
    ActionItem,
    LifecycleFlowPrioritizationResult,
    PrioritySlice,
    UnknownClassification,
)


def write_summary_excel(result: LifecycleFlowPrioritizationResult, output_dir: Path | str) -> Path:
    """Persist the prioritisation result as an Excel workbook."""

    path = Path(output_dir) / "upstream_priority.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    upstream_ws = workbook.active
    upstream_ws.title = "Upstream"
    _populate_priority_sheet(
        upstream_ws,
        chain(result.raw_materials, result.energy, result.auxiliaries),
        include_downstream_fields=False,
    )

    downstream_ws = workbook.create_sheet("Downstream")
    _populate_priority_sheet(
        downstream_ws,
        result.downstream_outputs,
        include_downstream_fields=True,
    )

    unknown_ws = workbook.create_sheet("Unknown")
    _populate_unknown_sheet(unknown_ws, result.unknown_classifications)

    actions_ws = workbook.create_sheet("Actions")
    _populate_actions_sheet(actions_ws, result.actions)

    notes_ws = workbook.create_sheet("Notes")
    _populate_single_column_sheet(notes_ws, "note", result.notes)

    metadata_ws = workbook.create_sheet("Metadata")
    _populate_metadata_sheet(metadata_ws, result.metadata)

    workbook.save(path)
    return path


def _populate_priority_sheet(
    ws,
    slices: Iterable[PrioritySlice],
    *,
    include_downstream_fields: bool,
) -> None:
    headers = [
        "dataset_uuid",
        "dataset_name",
        "unit_family",
        "flow_role",
        "flow_type",
        "exchange_name_zh",
        "exchange_name_en",
        "flow_name_zh",
        "flow_name_en",
        "flow_uuid",
        "reference_unit",
        "total_amount",
        "share_percent",
        "cumulative_percent",
        "reference_process_count",
    ]
    if include_downstream_fields:
        headers.extend(["downstream_path", "downstream_action"])
    _append_with_header(ws, headers)
    for slice_ in slices:
        row = [
            slice_.dataset_uuid,
            slice_.dataset_name,
            slice_.unit_family,
            slice_.flow_role,
            slice_.flow_type,
            slice_.exchange_name_zh,
            slice_.exchange_name_en,
            slice_.flow_name_zh,
            slice_.flow_name_en,
            slice_.flow_uuid,
            slice_.reference_unit,
            slice_.total_amount,
            _format_percent(slice_.share),
            _format_percent(slice_.cumulative_share),
            slice_.reference_process_count,
        ]
        if include_downstream_fields:
            row.extend([slice_.downstream_path, slice_.downstream_action])
        ws.append(row)
    _finalise_sheet(ws, headers)


def _populate_unknown_sheet(ws, entries: Iterable[UnknownClassification]) -> None:
    headers = ["exchange_name", "dataset_uuid", "reason"]
    _append_with_header(ws, headers)
    for entry in entries:
        ws.append(
            [
                entry.exchange_name,
                entry.dataset_uuid,
                entry.reason,
            ]
        )
    _finalise_sheet(ws, headers)


def _populate_actions_sheet(ws, entries: Iterable[ActionItem]) -> None:
    headers = ["priority", "type", "summary", "evidence"]
    _append_with_header(ws, headers)
    for action in entries:
        ws.append(
            [
                action.priority,
                action.type,
                action.summary,
                "; ".join(action.evidence),
            ]
        )
    _finalise_sheet(ws, headers)


def _populate_single_column_sheet(ws, header: str, lines: Iterable[str]) -> None:
    _append_with_header(ws, [header])
    for line in lines:
        ws.append([line])
    _finalise_sheet(ws, [header])


def _populate_metadata_sheet(ws, metadata: dict[str, object]) -> None:
    headers = ["field", "value"]
    _append_with_header(ws, headers)
    for key, value in metadata.items():
        ws.append([key, value])
    _finalise_sheet(ws, headers)


def _append_with_header(ws, headers: list[str]) -> None:
    ws.append(headers)


def _finalise_sheet(ws, headers: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def _format_percent(value: float | None) -> str | None:
    if value is None:
        return None
    return f"{value * 100:.4g}%"
