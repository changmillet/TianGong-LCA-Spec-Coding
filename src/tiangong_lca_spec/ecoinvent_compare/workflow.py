"""Workflow orchestration for the ecoinvent comparison CLI."""

from __future__ import annotations

import json
import subprocess
import sys
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping

from tiangong_lca_spec.core.config import Settings, get_settings
import time

from tiangong_lca_spec.core.exceptions import FlowSearchError
from tiangong_lca_spec.core.logging import get_logger
from tiangong_lca_spec.core.mcp_client import MCPToolClient
from tiangong_lca_spec.core.models import FlowCandidate, FlowQuery
from tiangong_lca_spec.flow_search.service import FlowSearchService
from tiangong_lca_spec.flow_search.validators import name_similarity_score

LOGGER = get_logger(__name__)

COMMON_NS = "http://lca.jrc.it/ILCD/Common"
FLOW_NS = "http://lca.jrc.it/ILCD/Flow"
PROCESS_NS = "http://lca.jrc.it/ILCD/Process"

FLOW_NAMESPACES = {"f": FLOW_NS, "common": COMMON_NS}
PROCESS_NAMESPACES = {"p": PROCESS_NS, "common": COMMON_NS}

NON_ELEMENTARY_TYPES = ("Product flow", "Waste flow")


def _text_or_none(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def _try_float(value: str | None) -> float | None:
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        try:
            return float(text.replace(",", ""))  # handle thousand separators if any
        except ValueError:
            return None


def _detect_git_revision(repo_root: Path) -> str | None:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=repo_root,
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:  # pragma: no cover - best effort
        return None
    return _text_or_none(result.stdout)


@dataclass(slots=True)
class FlowRecord:
    uuid: str
    name: str
    flow_type: str
    classification: list[dict[str, Any]]
    flow_property: str | None
    synonyms: list[str]
    version: str | None


@dataclass(slots=True)
class ProcessExchangeUsage:
    flow_uuid: str
    exchange_id: str | None
    direction: str | None
    mean_amount: float | None
    resulting_amount: float | None
    default_provider: str | None
    short_description: str | None
    general_comment: str | None
    unit_id: str | None
    property_id: str | None


@dataclass(slots=True)
class ProcessUsage:
    process_uuid: str
    process_name: str | None
    geography: str | None
    exchanges: list[ProcessExchangeUsage] = field(default_factory=list)


@dataclass(slots=True)
class FlowUsageSummary:
    flow: FlowRecord
    usage_count: int
    exchange_occurrences: int
    process_usages: list[ProcessUsage]

    def to_detail_entry(self) -> dict[str, Any]:
        return {
            "flow_uuid": self.flow.uuid,
            "flow_name": self.flow.name,
            "flow_type": self.flow.flow_type,
            "classification": self.flow.classification,
            "flow_property": self.flow.flow_property,
            "synonyms": self.flow.synonyms,
            "version": self.flow.version,
            "usage_count": self.usage_count,
            "exchange_occurrences": self.exchange_occurrences,
            "process_usages": [
                {
                    "process_uuid": proc.process_uuid,
                    "process_name": proc.process_name,
                    "geography": proc.geography,
                    "exchanges": [asdict(exchange) for exchange in proc.exchanges],
                }
                for proc in self.process_usages
            ],
        }

    def to_metadata(self) -> dict[str, Any]:
        return {
            "flow_uuid": self.flow.uuid,
            "flow_name": self.flow.name,
            "flow_type": self.flow.flow_type,
            "classification": self.flow.classification,
            "flow_property": self.flow.flow_property,
            "synonyms": self.flow.synonyms,
            "version": self.flow.version,
            "usage_count": self.usage_count,
            "exchange_occurrences": self.exchange_occurrences,
        }


@dataclass(slots=True)
class FlowUsageAccumulator:
    flow: FlowRecord
    process_usages: dict[str, ProcessUsage] = field(default_factory=dict)

    def add_exchange(self, process_uuid: str, process_name: str | None, geography: str | None, exchange: ProcessExchangeUsage) -> None:
        usage = self.process_usages.get(process_uuid)
        if usage is None:
            usage = ProcessUsage(
                process_uuid=process_uuid,
                process_name=process_name,
                geography=geography,
            )
            self.process_usages[process_uuid] = usage
        usage.exchanges.append(exchange)

    def to_summary(self) -> FlowUsageSummary:
        exchanges_total = sum(len(item.exchanges) for item in self.process_usages.values())
        return FlowUsageSummary(
            flow=self.flow,
            usage_count=len(self.process_usages),
            exchange_occurrences=exchanges_total,
            process_usages=list(self.process_usages.values()),
        )


@dataclass(slots=True)
class FlowSearchRecord:
    """JSON-serialisable container for the search stage."""

    metadata: dict[str, Any]
    matches: list[dict[str, Any]]
    unmatched: list[dict[str, Any]]
    errors: list[str]

    def to_dict(self) -> dict[str, Any]:
        return {
            **self.metadata,
            "matches": self.matches,
            "unmatched": self.unmatched,
            "errors": self.errors,
        }


@dataclass(slots=True)
class ProcessFetchConfig:
    enabled: bool = False
    server_name: str | None = None
    search_tool_name: str | None = "Search_Processes_Tool"
    crud_tool_name: str | None = "Database_CRUD_Tool"
    cache_path: Path | None = None
    dataset_cache_dir: Path | None = None
    max_search_results: int = 10


@dataclass(slots=True)
class RunConfig:
    flow_dir: Path
    process_dir: Path
    detail_path: Path
    search_output_path: Path
    final_output_path: Path
    detail_reference_path: Path | None = None
    min_usage: int = 1
    enable_search: bool = False
    flow_search_server: str | None = None
    flow_search_tool: str | None = None
    process_fetch: ProcessFetchConfig = field(default_factory=ProcessFetchConfig)
    excel_output_path: Path | None = None
    show_progress: bool = False
    retry_empty_matches: bool = False
    flow_dataset_cache_dir: Path | None = None


class _NullProgressStage:
    __slots__ = ()

    def advance(self, step: int = 1) -> None:  # pragma: no cover - trivial
        return

    def finish(self) -> None:  # pragma: no cover - trivial
        return


class _ProgressStage:
    __slots__ = ("_name", "_total", "_current", "_active")

    def __init__(self, name: str, total: int) -> None:
        self._name = name
        self._total = max(total, 1)
        self._current = 0
        self._active = True
        self._emit()

    def advance(self, step: int = 1) -> None:
        if not self._active:
            return
        self._current = min(self._total, self._current + max(step, 0))
        self._emit()

    def finish(self) -> None:
        if not self._active:
            return
        self._current = self._total
        self._emit(final=True)
        sys.stdout.write("\n")
        sys.stdout.flush()
        self._active = False

    def _emit(self, *, final: bool = False) -> None:
        percent = (self._current / self._total) * 100 if self._total else 100
        message = f"[{self._name}] {self._current}/{self._total} ({percent:5.1f}%)"
        suffix = " done" if final else ""
        sys.stdout.write("\r" + message + suffix)
        sys.stdout.flush()


class _ProgressPrinter:
    __slots__ = ("_enabled",)

    def __init__(self, enabled: bool) -> None:
        self._enabled = enabled

    def start_stage(self, name: str, total: int) -> _ProgressStage | _NullProgressStage:
        if not self._enabled or total <= 0:
            return _NullProgressStage()
        return _ProgressStage(name, total)


class _ProcessReferenceCache:
    def __init__(self, path: Path) -> None:
        self._path = path
        self._data: dict[str, list[dict[str, Any]]] = self.read_entries(path)

    def get(self, flow_uuid: str) -> list[dict[str, Any]] | None:
        return self._data.get(flow_uuid)

    def store(self, flow_uuid: str, references: list[dict[str, Any]]) -> None:
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._data[flow_uuid] = references
        entry = json.dumps({"flow_uuid": flow_uuid, "references": references}, ensure_ascii=False)
        with self._path.open("a", encoding="utf-8") as handle:
            handle.write(entry + "\n")

    @staticmethod
    def read_entries(path: Path | None) -> dict[str, list[dict[str, Any]]]:
        data: dict[str, list[dict[str, Any]]] = {}
        if path is None or not path.exists():
            return data
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                stripped = line.strip()
                if not stripped:
                    continue
                try:
                    payload = json.loads(stripped)
                except json.JSONDecodeError:
                    LOGGER.warning("process_cache.decode_failed", line=stripped[:100])
                    continue
                flow_uuid = payload.get("flow_uuid")
                references = payload.get("references") or []
                if isinstance(flow_uuid, str):
                    data[flow_uuid] = references if isinstance(references, list) else []
        return data


class _ProcessReferenceFetcher:
    def __init__(self, config: ProcessFetchConfig, settings: Settings) -> None:
        self._config = config
        self._settings = settings
        self._client: MCPToolClient | None = None
        self._cache = _ProcessReferenceCache(config.cache_path) if config.cache_path else None
        self._dataset_cache_dir = config.dataset_cache_dir

    def __enter__(self) -> "_ProcessReferenceFetcher":
        if self._config.enabled:
            self._client = MCPToolClient(self._settings)
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        if self._client is not None:
            self._client.close()

    def fetch(self, flow_uuid: str, flow_name: str | None = None) -> tuple[list[dict[str, Any]], str | None]:
        if not self._config.enabled:
            LOGGER.debug("process_fetcher.disabled", flow_uuid=flow_uuid)
            return [], "process_fetch_disabled"
        if self._client is None:
            LOGGER.debug("process_fetcher.client_missing", flow_uuid=flow_uuid)
            return [], "process_fetch_client_missing"
        if self._config.server_name is None:
            LOGGER.warning("process_fetcher.server_missing", flow_uuid=flow_uuid)
            return [], "process_fetch_server_missing"
        if self._cache:
            cached = self._cache.get(flow_uuid)
            if cached:
                return cached, None
        candidates, search_reason = self._search_processes(flow_uuid, flow_name)
        if search_reason:
            return [], search_reason
        if not candidates:
            return [], "process_search_empty"
        references: list[dict[str, Any]] = []
        dataset_missing = False
        missing_reference = False
        dataset_errors = False
        for process_id in candidates:
            dataset, dataset_reason = self._load_process_dataset(process_id)
            if dataset is None:
                dataset_missing = True
                if dataset_reason:
                    dataset_errors = True
                continue
            if not _process_references_flow(dataset, flow_uuid):
                missing_reference = True
                continue
            summary = _summarize_process_dataset(dataset)
            if summary:
                references.append(summary)
        if references:
            if self._cache is not None:
                self._cache.store(flow_uuid, references)
            LOGGER.debug("process_fetcher.response", flow_uuid=flow_uuid, reference_count=len(references))
            return references, None
        reason = None
        if dataset_missing and not missing_reference:
            reason = "process_dataset_missing" if not dataset_errors else "process_dataset_error"
        elif missing_reference:
            reason = "process_reference_missing"
        else:
            reason = "process_reference_unknown"
        LOGGER.debug("process_fetcher.response", flow_uuid=flow_uuid, reference_count=0, reason=reason)
        return [], reason

    def _search_processes(self, flow_uuid: str, flow_name: str | None) -> tuple[list[str], str | None]:
        assert self._client is not None
        tool_name = self._config.search_tool_name
        if not tool_name:
            return [], "process_search_tool_missing"
        query_tokens = [f"flow_uuid:{flow_uuid}"]
        if flow_name:
            query_tokens.append(flow_name)
        arguments = {"query": " ".join(query_tokens)}
        try:
            payload = self._client.invoke_json_tool(
                self._config.server_name,
                tool_name,
                arguments,
            )
        except Exception as exc:  # pragma: no cover
            LOGGER.error(
                "process_fetcher.search_failed",
                flow_uuid=flow_uuid,
                error=str(exc),
                tool=tool_name,
            )
            return [], "process_search_failed"
        identifiers = _extract_process_ids(payload)
        if self._config.max_search_results > 0:
            return identifiers[: self._config.max_search_results], None
        return identifiers, None

    def _load_process_dataset(self, process_id: str) -> tuple[dict[str, Any] | None, str | None]:
        assert self._client is not None
        tool_name = self._config.crud_tool_name
        if not tool_name:
            return None, "process_crud_tool_missing"
        cached = self._load_cached_dataset(process_id)
        if cached is not None:
            return cached, None
        arguments = {
            "operation": "select",
            "table": "processes",
            "id": process_id,
            "limit": 1,
        }
        try:
            payload = self._client.invoke_json_tool(
                self._config.server_name,
                tool_name,
                arguments,
            )
        except Exception as exc:  # pragma: no cover
            LOGGER.error(
                "process_fetcher.invoke_failed",
                flow_uuid=process_id,
                error=str(exc),
                tool=tool_name,
                server=self._config.server_name,
            )
            return None, "process_crud_failed"
        dataset = _extract_process_dataset(payload)
        if dataset is None:
            return None, "process_dataset_invalid"
        self._cache_dataset(process_id, dataset)
        return dataset, None

    def _load_cached_dataset(self, process_id: str) -> dict[str, Any] | None:
        if not self._dataset_cache_dir:
            return None
        path = self._dataset_cache_dir / f"{process_id}.json"
        if not path.exists():
            return None
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:  # pragma: no cover - best effort
            return None

    def _cache_dataset(self, process_id: str, dataset: dict[str, Any]) -> None:
        if not self._dataset_cache_dir:
            return
        try:
            self._dataset_cache_dir.mkdir(parents=True, exist_ok=True)
            path = self._dataset_cache_dir / f"{process_id}.json"
            path.write_text(json.dumps(dataset, ensure_ascii=False), encoding="utf-8")
        except Exception:  # pragma: no cover - best effort
            LOGGER.warning("process_fetcher.cache_dataset_failed", process_id=process_id)


class _FlowDatasetFetcher:
    def __init__(self, server_name: str | None, crud_tool_name: str | None, cache_dir: Path | None, settings: Settings) -> None:
        self._server_name = server_name
        self._crud_tool_name = crud_tool_name or "Database_CRUD_Tool"
        self._cache_dir = cache_dir
        self._settings = settings
        self._client: MCPToolClient | None = None

    def __enter__(self) -> "_FlowDatasetFetcher":
        if self._server_name:
            self._client = MCPToolClient(self._settings)
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        if self._client:
            self._client.close()

    def fetch_type(self, flow_uuid: str) -> tuple[str | None, str | None]:
        if not flow_uuid:
            return None, "flow_uuid_missing"
        cached = self._load_cached(flow_uuid)
        if cached is not None:
            return cached, None
        if self._client is None or self._server_name is None:
            return None, "flow_dataset_fetch_disabled"
        arguments = {
            "operation": "select",
            "table": "flows",
            "id": flow_uuid,
            "limit": 1,
        }
        try:
            payload = self._client.invoke_json_tool(self._server_name, self._crud_tool_name, arguments)
        except Exception as exc:  # pragma: no cover
            LOGGER.warning("flow_dataset.fetch_failed", flow_uuid=flow_uuid, error=str(exc))
            return None, "flow_dataset_fetch_failed"
        dataset = _extract_flow_dataset(payload)
        if dataset is None:
            return None, "flow_dataset_invalid"
        flow_type = _extract_flow_type(dataset)
        self._cache_dataset(flow_uuid, dataset)
        return flow_type, None

    def _cache_dataset(self, flow_uuid: str, dataset: dict[str, Any]) -> None:
        if not self._cache_dir:
            return
        try:
            self._cache_dir.mkdir(parents=True, exist_ok=True)
            path = self._cache_dir / f"{flow_uuid}.json"
            path.write_text(json.dumps(dataset, ensure_ascii=False), encoding="utf-8")
        except Exception:  # pragma: no cover
            LOGGER.warning("flow_dataset.cache_failed", flow_uuid=flow_uuid)

    def _load_cached(self, flow_uuid: str) -> str | None:
        if not self._cache_dir:
            return None
        path = self._cache_dir / f"{flow_uuid}.json"
        if not path.exists():
            return None
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:  # pragma: no cover
            return None
        flow_type = _extract_flow_type(payload)
        return flow_type


def run_workflow(config: RunConfig) -> dict[str, Any]:
    """Entry point used by the CLI."""
    LOGGER.info(
        "ecoinvent_compare.start",
        flow_dir=str(config.flow_dir),
        process_dir=str(config.process_dir),
        min_usage=config.min_usage,
    )
    settings = get_settings()
    progress = _ProgressPrinter(config.show_progress)
    process_cache_path = config.process_fetch.cache_path or (settings.cache_dir / "tiangong_flow_process.jsonl")
    summaries, stats = _build_flow_usage(config.flow_dir, config.process_dir, config.min_usage)
    _write_jsonl(config.detail_path, (summary.to_detail_entry() for summary in summaries))
    LOGGER.info("flow_usage.details_written", path=str(config.detail_path), summary_count=len(summaries))
    flow_stage = progress.start_stage("Flow search", len(summaries))
    _run_flow_search(
        summaries,
        config.search_output_path,
        enable_search=config.enable_search,
        server_override=config.flow_search_server,
        tool_override=config.flow_search_tool,
        retry_empty_matches=config.retry_empty_matches,
        flow_dataset_cache_dir=config.flow_dataset_cache_dir,
        progress_stage=flow_stage,
    )
    LOGGER.info("flow_search.results_written", path=str(config.search_output_path))
    _sync_tiangong_flow_cache(config.search_output_path, process_cache_path)
    LOGGER.info("process_cache.synced", path=str(process_cache_path))
    search_records = list(_read_jsonl(config.search_output_path))
    fetch_stage = progress.start_stage("Process fetch", len(search_records))
    payload = _assemble_final_report(
        config=config,
        stats=stats,
        settings=settings,
        process_cache_path=process_cache_path,
        records=search_records,
        progress_stage=fetch_stage,
    )
    if config.excel_output_path:
        _export_overview_excel(
            payload,
            config.search_output_path,
            process_cache_path,
            config.excel_output_path,
        )
        LOGGER.info("excel_export.completed", path=str(config.excel_output_path))
    return payload


def _build_flow_usage(flow_dir: Path, process_dir: Path, min_usage: int) -> tuple[list[FlowUsageSummary], dict[str, Any]]:
    if not flow_dir.is_dir():
        raise FileNotFoundError(f"Flow directory not found: {flow_dir}")
    if not process_dir.is_dir():
        raise FileNotFoundError(f"Process directory not found: {process_dir}")
    flow_records = _index_flows(flow_dir)
    accumulators: dict[str, FlowUsageAccumulator] = {}
    processed_process_files = 0
    for process_file in sorted(process_dir.glob("*.xml")):
        processed_process_files += 1
        try:
            tree = ET.parse(process_file)
        except ET.ParseError as exc:
            LOGGER.warning("process.parse_failed", file=str(process_file), error=str(exc))
            continue
        root = tree.getroot()
        process_uuid = _text_or_none(root.findtext(".//common:UUID", namespaces=PROCESS_NAMESPACES))
        if not process_uuid:
            continue
        process_name = _text_or_none(root.findtext(".//p:name/p:baseName", namespaces=PROCESS_NAMESPACES))
        geography = _extract_location(root)
        exchanges_parent = root.find("p:exchanges", namespaces=PROCESS_NAMESPACES)
        if exchanges_parent is None:
            continue
        for exchange in exchanges_parent.findall("p:exchange", namespaces=PROCESS_NAMESPACES):
            reference = exchange.find("p:referenceToFlowDataSet", namespaces=PROCESS_NAMESPACES)
            if reference is None:
                continue
            flow_uuid = reference.attrib.get("refObjectId")
            if not flow_uuid:
                continue
            record = flow_records.get(flow_uuid)
            if record is None:
                continue
            accumulator = accumulators.setdefault(flow_uuid, FlowUsageAccumulator(flow=record))
            default_provider = exchange.attrib.get("{http://openlca.org/ilcd-extensions}defaultProvider") or exchange.attrib.get("olca:defaultProvider")
            unit_id = exchange.attrib.get("{http://openlca.org/ilcd-extensions}unitId") or exchange.attrib.get("olca:unitId")
            property_id = exchange.attrib.get("{http://openlca.org/ilcd-extensions}propertyId") or exchange.attrib.get("olca:propertyId")
            accumulator.add_exchange(
                process_uuid=process_uuid,
                process_name=process_name,
                geography=geography,
                exchange=ProcessExchangeUsage(
                    flow_uuid=flow_uuid,
                    exchange_id=exchange.attrib.get("dataSetInternalID"),
                    direction=_text_or_none(exchange.findtext("p:exchangeDirection", namespaces=PROCESS_NAMESPACES)),
                    mean_amount=_try_float(exchange.findtext("p:meanAmount", namespaces=PROCESS_NAMESPACES)),
                    resulting_amount=_try_float(exchange.findtext("p:resultingAmount", namespaces=PROCESS_NAMESPACES)),
                    default_provider=_text_or_none(default_provider),
                    short_description=_text_or_none(reference.findtext("common:shortDescription", namespaces=PROCESS_NAMESPACES)),
                    general_comment=_text_or_none(exchange.findtext("common:generalComment", namespaces=PROCESS_NAMESPACES)),
                    unit_id=_text_or_none(unit_id),
                    property_id=_text_or_none(property_id),
                ),
            )
    summaries = [acc.to_summary() for acc in accumulators.values()]
    filtered = [summary for summary in summaries if summary.usage_count >= max(1, min_usage)]
    filtered.sort(key=lambda summary: (-summary.usage_count, summary.flow.name.lower()))
    stats = {
        "flows_indexed": len(flow_records),
        "process_files_scanned": processed_process_files,
        "flows_with_usage": len(summaries),
        "flows_after_filter": len(filtered),
        "min_usage": max(1, min_usage),
    }
    LOGGER.info("flow_usage.completed", stats=stats)
    return filtered, stats


def _index_flows(flow_dir: Path) -> dict[str, FlowRecord]:
    flow_records: dict[str, FlowRecord] = {}
    total_files = 0
    included = 0
    for flow_file in sorted(flow_dir.glob("*.xml")):
        total_files += 1
        try:
            tree = ET.parse(flow_file)
        except ET.ParseError as exc:
            LOGGER.warning("flow.parse_failed", file=str(flow_file), error=str(exc))
            continue
        root = tree.getroot()
        flow_type = _text_or_none(root.findtext(".//f:typeOfDataSet", namespaces=FLOW_NAMESPACES))
        if flow_type not in NON_ELEMENTARY_TYPES:
            continue
        uuid = _text_or_none(root.findtext(".//common:UUID", namespaces=FLOW_NAMESPACES))
        name = _text_or_none(root.findtext(".//f:name/f:baseName", namespaces=FLOW_NAMESPACES))
        if not uuid or not name:
            continue
        classification = _extract_classification(root)
        flow_property = _extract_flow_property(root)
        synonyms = _extract_synonyms(root)
        version = _text_or_none(root.findtext(".//f:publicationAndOwnership/common:dataSetVersion", namespaces=FLOW_NAMESPACES))
        flow_records[uuid] = FlowRecord(
            uuid=uuid,
            name=name,
            flow_type=flow_type,
            classification=classification,
            flow_property=flow_property,
            synonyms=synonyms,
            version=version,
        )
        included += 1
    LOGGER.info("flow_index.completed", total_files=total_files, included=included)
    return flow_records


def _extract_classification(root: ET.Element) -> list[dict[str, Any]]:
    categories: list[dict[str, Any]] = []
    for node in root.findall(".//common:elementaryFlowCategorization/common:category", namespaces=FLOW_NAMESPACES):
        categories.append(
            {
                "level": _parse_int(node.attrib.get("level")),
                "description": _text_or_none(node.text),
            }
        )
    for node in root.findall(".//common:classification/common:class", namespaces=FLOW_NAMESPACES):
        categories.append(
            {
                "level": _parse_int(node.attrib.get("level")),
                "description": _text_or_none(node.text),
            }
        )
    return [cat for cat in categories if cat["description"]]


def _extract_synonyms(root: ET.Element) -> list[str]:
    synonyms: list[str] = []
    for node in root.findall(".//common:synonyms", namespaces=FLOW_NAMESPACES):
        text = _text_or_none(node.text)
        if text:
            synonyms.append(text)
    return synonyms


def _extract_flow_property(root: ET.Element) -> str | None:
    for node in root.findall(".//f:flowProperties/f:flowProperty//common:shortDescription", namespaces=FLOW_NAMESPACES):
        text = _text_or_none(node.text)
        if text:
            return text
    return None


def _parse_int(value: str | None) -> int | str | None:
    if value is None:
        return None
    try:
        return int(value)
    except ValueError:
        return value


def _extract_location(root: ET.Element) -> str | None:
    node = root.find(".//p:geography/p:locationOfOperationSupplyOrProduction", namespaces=PROCESS_NAMESPACES)
    if node is None:
        return None
    location = node.attrib.get("location")
    return _text_or_none(location)


def _write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False))
            handle.write("\n")


def _load_flow_search_map(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    records: dict[str, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            stripped = line.strip()
            if not stripped:
                continue
            record = json.loads(stripped)
            flow_uuid = record.get("flow_uuid")
            if flow_uuid:
                records[flow_uuid] = record
    return records


def _write_flow_search_records(path: Path, summaries: list[FlowUsageSummary], record_map: dict[str, dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for summary in summaries:
            key = summary.flow.uuid
            record = record_map.get(key)
            if record is None:
                record_obj = FlowSearchRecord(metadata=summary.to_metadata(), matches=[], unmatched=[], errors=["Flow search pending"])
                record = record_obj.to_dict()
                record_map[key] = record
            handle.write(json.dumps(record, ensure_ascii=False))
            handle.write("\n")


def _sync_tiangong_flow_cache(search_output_path: Path, cache_path: Path) -> None:
    cache: dict[str, list[dict[str, Any]]] = {}
    changed = not cache_path.exists()
    if cache_path.exists():
        with cache_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                stripped = line.strip()
                if not stripped:
                    continue
                payload = json.loads(stripped)
                flow_uuid = payload.get("flow_uuid")
                references = payload.get("references") or []
                if isinstance(flow_uuid, str):
                    cache[flow_uuid] = references
    if not search_output_path.exists():
        if changed:
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            cache_path.write_text("", encoding="utf-8")
        return
    with search_output_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            stripped = line.strip()
            if not stripped:
                continue
            record = json.loads(stripped)
            for match in record.get("matches") or []:
                flow_uuid = match.get("uuid") or match.get("flow_uuid")
                if not flow_uuid:
                    continue
                if flow_uuid not in cache:
                    cache[flow_uuid] = []
                    changed = True
    if not changed:
        return
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with cache_path.open("w", encoding="utf-8") as handle:
        for flow_uuid, references in cache.items():
            handle.write(json.dumps({"flow_uuid": flow_uuid, "references": references}, ensure_ascii=False))
            handle.write("\n")


def _export_overview_excel(
    final_payload: dict[str, Any],
    search_results_path: Path,
    process_cache_path: Path,
    output_path: Path,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        LOGGER.warning("excel_export.openpyxl_missing")
        return

    summary_block = final_payload.get("summary", {})
    results = final_payload.get("results") or []
    detail_path_str = (final_payload.get("source") or {}).get("flow_usage_details")
    detail_path = Path(detail_path_str) if detail_path_str else None

    search_records = _load_jsonl_records(search_results_path)
    matched_tiangong_ids: set[str] = set()
    matched_tiangong_product_ids: set[str] = set()
    ecoinvent_product_flow_ids: set[str] = set()
    for record in search_records:
        if (record.get("flow_type") or "").strip().lower() == "product flow":
            if isinstance(record.get("flow_uuid"), str):
                ecoinvent_product_flow_ids.add(record["flow_uuid"])
        for match in record.get("matches") or []:
            match_uuid = match.get("uuid") or match.get("flow_uuid")
            if match_uuid:
                matched_tiangong_ids.add(match_uuid)
                if (match.get("flow_type_tiangong") or "").strip().lower() == "product flow":
                    matched_tiangong_product_ids.add(match_uuid)

    process_cache = _ProcessReferenceCache.read_entries(process_cache_path)
    tiangong_with_refs = {uuid for uuid, refs in process_cache.items() if refs}
    reference_total = sum(len(refs) for refs in process_cache.values())

    unique_process_usage_ids: set[str] = set()
    if detail_path and detail_path.exists():
        try:
            with detail_path.open("r", encoding="utf-8") as handle:
                for line in handle:
                    stripped = line.strip()
                    if not stripped:
                        continue
                    try:
                        entry = json.loads(stripped)
                    except json.JSONDecodeError:
                        continue
                    for usage in entry.get("process_usages") or []:
                        proc_uuid = usage.get("process_uuid")
                        if isinstance(proc_uuid, str):
                            unique_process_usage_ids.add(proc_uuid)
        except Exception:  # pragma: no cover - best effort
            LOGGER.warning("excel_export.process_usage_load_failed", path=str(detail_path))

    wb = Workbook()
    ws_stats = wb.active
    ws_stats.title = "stats"
    ws_stats.append(["Metric", "Value"])
    ws_stats.append(["Unique ecoinvent product flows (retained)", len(ecoinvent_product_flow_ids)])
    ws_stats.append(["Total ecoinvent process files", summary_block.get("process_files_scanned")])
    ws_stats.append(["Processes referencing retained flows", len(unique_process_usage_ids)])
    ws_stats.append(["Flows with Tiangong matches", len([r for r in search_records if r.get("matches")])])
    ws_stats.append(["Unique matched Tiangong flows", len(matched_tiangong_ids)])
    ws_stats.append(["Matched Tiangong product flows", len(matched_tiangong_product_ids)])
    ws_stats.append(["Tiangong flows tracked in process cache", len(process_cache)])
    ws_stats.append(["Tiangong flows with process references", len(tiangong_with_refs)])
    ws_stats.append(["Total Tiangong process references", reference_total])

    ws_search = wb.create_sheet("flow_search_matches")
    ws_search.append(
        [
            "ecoinvent_flow_uuid",
            "ecoinvent_flow_name",
            "flow_type",
            "usage_count",
            "exchange_occurrences",
            "tiangong_flow_uuid",
            "tiangong_flow_name",
            "similarity",
            "treatment_routes",
            "mix_location",
            "flow_properties",
            "geography",
            "classification",
            "notes",
        ]
    )
    for record in search_records:
        base_row = [
            record.get("flow_uuid"),
            record.get("flow_name"),
            record.get("flow_type"),
            record.get("usage_count"),
            record.get("exchange_occurrences"),
        ]
        notes = "|".join(record.get("notes") or [])
        matches = record.get("matches") or [{}]
        for match in matches:
            if not match:
                continue
            ws_search.append(
                base_row
                + [
                    match.get("uuid") or match.get("flow_uuid"),
                    match.get("flow_name") or match.get("base_name"),
                    match.get("similarity"),
                    match.get("treatment_standards_routes"),
                    match.get("mix_and_location_types"),
                    match.get("flow_properties"),
                    _format_geography_text(match.get("geography")),
                    _format_classification_text(match.get("classification")),
                    notes,
                ]
            )

    ws_process = wb.create_sheet("tiangong_process_refs")
    ws_process.append(["tiangong_flow_uuid", "reference_count", "process_uuid", "process_name", "type_of_data_set", "geography"])
    for flow_uuid, references in process_cache.items():
        if not references:
            ws_process.append([flow_uuid, 0, "", "", "", ""])
            continue
        count = len(references)
        for ref in references:
            ws_process.append(
                [
                    flow_uuid,
                    count,
                    ref.get("process_uuid"),
                    ref.get("process_name"),
                    ref.get("type_of_data_set"),
                    ref.get("geography"),
                ]
            )

    ws_final = wb.create_sheet("final_similarity")
    ws_final.append(
        [
            "flow_uuid",
            "flow_name",
            "flow_type",
            "usage_count",
            "exchange_occurrences",
            "tiangong_flow_uuid",
            "tiangong_flow_name",
            "similarity",
            "treatment_routes",
            "mix_location",
            "flow_properties",
            "geography",
            "process_reference_count",
            "process_reference_note",
            "notes",
        ]
    )
    for result in results:
        notes = "|".join(result.get("notes") or [])
        base_row = [
            result.get("flow_uuid"),
            result.get("flow_name"),
            result.get("flow_type"),
            result.get("usage_count"),
            result.get("exchange_occurrences"),
        ]
        matches = result.get("tiangong_matches") or [{}]
        for match in matches:
            ws_final.append(
                base_row
                + [
                    match.get("flow_uuid"),
                    match.get("flow_name"),
                    match.get("similarity"),
                    match.get("treatment_standards_routes"),
                    match.get("mix_and_location_types"),
                    match.get("flow_properties"),
                    match.get("geography"),
                    match.get("process_reference_count"),
                    match.get("process_reference_note"),
                    notes,
                ]
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _load_jsonl_records(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if not path.exists():
        return records
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            stripped = line.strip()
            if stripped:
                records.append(json.loads(stripped))
    return records


def _format_classification_text(classification: Any) -> str:
    if not isinstance(classification, list):
        return ""
    parts = []
    for item in classification:
        if not isinstance(item, dict):
            continue
        level = item.get("@level") or item.get("level") or "?"
        text = item.get("#text") or item.get("description")
        if text:
            parts.append(f"{level}: {text}")
    return " | ".join(parts)


def _run_flow_search(
    summaries: list[FlowUsageSummary],
    output_path: Path,
    *,
    enable_search: bool,
    server_override: str | None,
    tool_override: str | None,
    retry_empty_matches: bool = False,
    flow_dataset_cache_dir: Path | None = None,
    progress_stage: _ProgressStage | _NullProgressStage | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    existing_records = _load_flow_search_map(output_path)
    service: FlowSearchService | None = None
    if enable_search:
        service = FlowSearchService(server_name=server_override, tool_name=tool_override)
    settings = get_settings()
    flow_dataset_fetcher = _FlowDatasetFetcher(server_override or settings.flow_search_service_name, tool_override or "Database_CRUD_Tool", flow_dataset_cache_dir, settings)
    flow_dataset_cm = flow_dataset_fetcher if flow_dataset_cache_dir else None
    progress_stage = progress_stage or _NullProgressStage()
    processed_since_flush = 0
    updated = 0
    try:
        if flow_dataset_cm:
            flow_dataset_cm.__enter__()
        for summary in summaries:
            flow_uuid = summary.flow.uuid
            record = existing_records.get(flow_uuid)
            if record:
                record.update(summary.to_metadata())
                matches = record.get("matches") or []
                has_matches = bool(matches)
                if has_matches:
                    if flow_dataset_cm:
                        _populate_flow_dataset_types(record, flow_dataset_fetcher)
                    existing_records[flow_uuid] = record
                    progress_stage.advance()
                    continue
                if not retry_empty_matches:
                    record["flow_name"] = record.get("flow_name") or summary.flow.name
                    existing_records[flow_uuid] = record
                    progress_stage.advance()
                    continue
            while True:
                try:
                    if not enable_search or service is None:
                        record = FlowSearchRecord(
                            metadata=summary.to_metadata(),
                            matches=[],
                            unmatched=[],
                            errors=["Flow search disabled"],
                        ).to_dict()
                        record["flow_search_completed"] = True
                        record["flow_name"] = summary.flow.name
                        existing_records[flow_uuid] = record
                        break
                    record = _build_search_record(summary, service, flow_dataset_fetcher=flow_dataset_fetcher if flow_dataset_cm else None)
                    record["flow_name"] = record.get("flow_name") or summary.flow.name
                    existing_records[flow_uuid] = record
                    updated += 1
                    break
                except FlowSearchError as exc:
                    LOGGER.warning("flow_search.retry_wait", flow_uuid=flow_uuid, error=str(exc))
                    time.sleep(60)
                    continue
            processed_since_flush += 1
            if processed_since_flush >= 50:
                _write_flow_search_records(output_path, summaries, existing_records)
                processed_since_flush = 0
            progress_stage.advance()
        _write_flow_search_records(output_path, summaries, existing_records)
        LOGGER.info("flow_search.completed", updated=updated, total=len(summaries))
    except KeyboardInterrupt:
        LOGGER.warning("flow_search.interrupted", message="中断，请继续执行")
        _write_flow_search_records(output_path, summaries, existing_records)
        raise
    except Exception:
        LOGGER.error("flow_search.interrupted", message="中断，请继续执行")
        _write_flow_search_records(output_path, summaries, existing_records)
        raise
    finally:
        if service is not None:
            service.close()
        if flow_dataset_cm:
            flow_dataset_cm.__exit__(None, None, None)
        progress_stage.finish()


def _build_search_record(
    summary: FlowUsageSummary,
    service: FlowSearchService,
    *,
    flow_dataset_fetcher: "_FlowDatasetFetcher" | None = None,
) -> dict[str, Any]:
    metadata = summary.to_metadata()
    matches: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    errors: list[str] = []
    notes: list[str] = []
    query = FlowQuery(
        exchange_name=summary.flow.name,
        description=_build_query_description(summary),
    )
    try:
        candidates, filtered = service.lookup(query)
        serialized_matches = [_serialize_candidate(query, candidate) for candidate in candidates]
        matches = serialized_matches
        if len(matches) > 1:
            matches = matches[:1]
        if flow_dataset_fetcher:
            _populate_flow_dataset_types({"matches": matches}, flow_dataset_fetcher, notes)
        for match in matches:
            match["flow_name"] = _compose_candidate_name(match)
        unmatched = [{"base_name": item.base_name, "process_name": item.process_name} for item in filtered]
        if not matches and not unmatched:
            errors.append("Flow search returned no candidates")
        elif not matches and unmatched:
            snippet = ", ".join(item["base_name"] for item in unmatched[:5] if item.get("base_name"))
            if snippet:
                notes.append(f"low_similarity_candidates:{snippet}")
    except FlowSearchError as exc:
        LOGGER.error(
            "flow_search.failed",
            flow_uuid=summary.flow.uuid,
            flow_name=summary.flow.name,
            error=str(exc),
        )
        message = str(exc)
        if "timeout" in message.lower():
            errors.append(f"Flow search timeout: {message}")
        else:
            errors.append(f"Flow search error: {message}")
    except Exception as exc:  # pylint: disable=broad-except
        LOGGER.error(
            "flow_search.failed",
            flow_uuid=summary.flow.uuid,
            flow_name=summary.flow.name,
            error=str(exc),
        )
        errors.append(f"Flow search error: {exc}")
    record = FlowSearchRecord(metadata=metadata, matches=matches, unmatched=unmatched, errors=errors).to_dict()
    if notes:
        record["notes"] = list(set(record.get("notes") or []) | set(notes))
    record["flow_search_completed"] = True
    return record


def _build_query_description(summary: FlowUsageSummary) -> str | None:
    if summary.flow.classification:
        for item in reversed(summary.flow.classification):
            description = _text_or_none(item.get("description"))
            if description:
                return description
    if summary.flow.synonyms:
        return ", ".join(summary.flow.synonyms[:3])
    if summary.flow.flow_property:
        return summary.flow.flow_property
    return None


def _serialize_candidate(query: FlowQuery, candidate: FlowCandidate) -> dict[str, Any]:
    candidate_dict = asdict(candidate)
    similarity = name_similarity_score(query, candidate_dict)
    candidate_dict["similarity"] = similarity
    return candidate_dict


def _format_geography_text(geography: Any) -> str | None:
    if geography is None:
        return None
    if isinstance(geography, dict):
        return geography.get("@location") or geography.get("location") or geography.get("description") or geography.get("code")
    if isinstance(geography, str):
        return geography
    return None


def _compose_candidate_name(match: Mapping[str, Any]) -> str:
    base_name = _text_or_none(match.get("base_name") or match.get("flow_name")) or "-"
    treatment = _text_or_none(match.get("treatment_standards_routes")) or "-"
    location = _text_or_none(_format_geography_text(match.get("geography"))) or "-"
    flow_property = _text_or_none(match.get("flow_properties")) or "-"
    return "; ".join((base_name, treatment, location, flow_property))


def _populate_flow_dataset_types(record: dict[str, Any], fetcher: "_FlowDatasetFetcher", notes: list[str] | None = None) -> None:
    matches = record.get("matches") or []
    if not matches:
        return
    for match in matches:
        flow_uuid = match.get("uuid") or match.get("flow_uuid")
        flow_type_tg, dataset_reason = fetcher.fetch_type(flow_uuid)
        if flow_type_tg:
            match["flow_type_tiangong"] = flow_type_tg
        if dataset_reason and notes is not None:
            notes.append(f"flow_dataset:{dataset_reason}")


def _extract_process_ids(payload: Any) -> list[str]:
    if payload is None:
        return []
    if isinstance(payload, str):
        try:
            payload = json.loads(payload)
        except json.JSONDecodeError:
            return []
    candidates: list[Any]
    if isinstance(payload, list):
        candidates = payload
    elif isinstance(payload, dict):
        for key in ("results", "processes", "data"):
            value = payload.get(key)
            if isinstance(value, list):
                candidates = value
                break
        else:
            candidates = [payload]
    else:
        return []
    identifiers: list[str] = []
    for item in candidates:
        if not isinstance(item, dict):
            continue
        for key in ("process_uuid", "uuid", "id"):
            value = item.get(key)
            if isinstance(value, str):
                identifiers.append(value)
                break
        else:
            dataset = item.get("processDataSet")
            if isinstance(dataset, dict):
                uuid = dataset.get("processInformation", {}).get("dataSetInformation", {}).get("common:UUID")
                if isinstance(uuid, str):
                    identifiers.append(uuid)
    return identifiers


def _extract_process_dataset(payload: Any) -> dict[str, Any] | None:
    data = payload
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except json.JSONDecodeError:
            return None
    if isinstance(data, dict):
        entries = data.get("data")
        entry: Any
        if isinstance(entries, list) and entries:
            entry = entries[0]
        else:
            entry = data
        if isinstance(entry, dict):
            dataset = entry.get("json_ordered") or entry.get("json")
            if isinstance(dataset, str):
                try:
                    dataset = json.loads(dataset)
                except json.JSONDecodeError:
                    return None
            if isinstance(dataset, dict):
                return dataset
    elif isinstance(data, list) and data:
        entry = data[0]
        if isinstance(entry, dict):
            dataset = entry.get("json_ordered") or entry.get("json")
            if isinstance(dataset, str):
                try:
                    dataset = json.loads(dataset)
                except json.JSONDecodeError:
                    return None
            if isinstance(dataset, dict):
                return dataset
    return None


def _extract_flow_dataset(payload: Any) -> dict[str, Any] | None:
    data = payload
    if isinstance(data, str):
        try:
            data = json.loads(data)
        except json.JSONDecodeError:
            return None
    if isinstance(data, dict):
        entries = data.get("data")
        entry: Any
        if isinstance(entries, list) and entries:
            entry = entries[0]
        else:
            entry = data
        if isinstance(entry, dict):
            dataset = entry.get("json") or entry.get("json_ordered") or entry.get("jsonOrdered")
            if isinstance(dataset, str):
                try:
                    dataset = json.loads(dataset)
                except json.JSONDecodeError:
                    return None
            if isinstance(dataset, dict):
                return dataset
    return None


def _extract_flow_type(dataset: dict[str, Any]) -> str | None:
    root = dataset.get("flowDataSet") or dataset
    if not isinstance(root, dict):
        return None
    modelling = root.get("modellingAndValidation", {})
    if isinstance(modelling, dict):
        lci = modelling.get("LCIMethod") or modelling.get("LCIMethodAndAllocation")
        if isinstance(lci, dict):
            flow_type = lci.get("typeOfDataSet")
            if isinstance(flow_type, str):
                return _text_or_none(flow_type)
    return None


def _process_references_flow(dataset: dict[str, Any], flow_uuid: str) -> bool:
    root = dataset.get("processDataSet") or dataset
    if not isinstance(root, dict):
        return False
    process_info = root.get("processInformation", {})
    quant_ref = process_info.get("quantitativeReference")
    if isinstance(quant_ref, dict):
        reference = quant_ref.get("referenceToReferenceFlow")
        if isinstance(reference, dict):
            ref_uuid = reference.get("@refObjectId") or reference.get("common:UUID")
            if ref_uuid == flow_uuid:
                return True
    exchanges_block = root.get("exchanges")
    for exchange in _iter_exchange_entries(exchanges_block):
        ref = exchange.get("referenceToFlowDataSet")
        if isinstance(ref, dict) and ref.get("@refObjectId") == flow_uuid:
            return True
    return False


def _iter_exchange_entries(block: Any) -> Iterator[dict[str, Any]]:
    if block is None:
        return
    if isinstance(block, list):
        for item in block:
            if isinstance(item, dict):
                yield item
    elif isinstance(block, dict):
        if "exchange" in block:
            yield from _iter_exchange_entries(block["exchange"])
        else:
            yield block


def _summarize_process_dataset(dataset: dict[str, Any]) -> dict[str, Any] | None:
    root = dataset.get("processDataSet") or dataset
    if not isinstance(root, dict):
        return None
    info = root.get("processInformation", {})
    data_info = info.get("dataSetInformation", {})
    uuid = _text_or_none(data_info.get("common:UUID"))
    if not uuid:
        return None
    name_block = data_info.get("name") or {}
    process_name = _preferred_language_text(name_block.get("baseName")) or _preferred_language_text(name_block)
    modelling = root.get("modellingAndValidation", {})
    type_of_data = None
    if isinstance(modelling, dict):
        lcim = modelling.get("LCIMethodAndAllocation")
        if isinstance(lcim, dict):
            type_of_data = lcim.get("typeOfDataSet")
    location = None
    geography = info.get("geography", {})
    if isinstance(geography, dict):
        location = geography.get("locationOfOperationSupplyOrProduction")
    geo_text = _format_geography_text(location)
    return {
        "process_uuid": uuid,
        "process_name": process_name or "unknown_process",
        "type_of_data_set": _text_or_none(type_of_data),
        "geography": _text_or_none(geo_text),
    }


ENGLISH_LANG_KEYS = ("en", "en-us", "en-gb", "english")
CHINESE_LANG_KEYS = ("zh-hans", "zh-cn", "zh", "简体中文")


def _preferred_language_text(value: Any) -> str | None:
    english = _find_language_text(value, ENGLISH_LANG_KEYS)
    if english:
        return english
    chinese = _find_language_text(value, CHINESE_LANG_KEYS)
    if chinese:
        return chinese
    return _first_text(value)


def _find_language_text(value: Any, language_keys: tuple[str, ...]) -> str | None:
    normalized = {_normalize_language_key(token) for token in language_keys}
    return _find_language_text_recursive(value, normalized)


def _find_language_text_recursive(value: Any, targets: set[str]) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, dict):
        lang = value.get("@xml:lang") or value.get("lang")
        text = value.get("#text") or value.get("text")
        if lang and isinstance(text, str):
            if _normalize_language_key(lang) in targets:
                stripped = text.strip()
                if stripped:
                    return stripped
        for child in value.values():
            result = _find_language_text_recursive(child, targets)
            if result:
                return result
    elif isinstance(value, list):
        for item in value:
            result = _find_language_text_recursive(item, targets)
            if result:
                return result
    return None


def _first_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, list):
        for item in value:
            result = _first_text(item)
            if result:
                return result
    if isinstance(value, dict):
        text = value.get("#text")
        if isinstance(text, str):
            stripped = text.strip()
            if stripped:
                return stripped
        for child in value.values():
            result = _first_text(child)
            if result:
                return result
    return None


def _normalize_language_key(token: str | None) -> str:
    return (token or "").strip().lower()


def _assemble_final_report(
    *,
    config: RunConfig,
    stats: dict[str, Any],
    settings: Settings,
    process_cache_path: Path,
    records: list[dict[str, Any]],
    progress_stage: _ProgressStage | _NullProgressStage | None = None,
) -> dict[str, Any]:
    detail_reference_path = config.detail_reference_path or config.detail_path
    search_records = records
    process_fetch_cfg = ProcessFetchConfig(
        enabled=config.process_fetch.enabled,
        server_name=config.process_fetch.server_name or settings.flow_search_service_name,
        search_tool_name=config.process_fetch.search_tool_name,
        crud_tool_name=config.process_fetch.crud_tool_name,
        cache_path=process_cache_path,
        dataset_cache_dir=config.process_fetch.dataset_cache_dir,
        max_search_results=config.process_fetch.max_search_results,
    )
    progress_stage = progress_stage or _NullProgressStage()
    results: list[dict[str, Any]] = []
    total_usage = 0
    with _ProcessReferenceFetcher(process_fetch_cfg, settings) as fetcher:
        for record in search_records:
            usage_count = int(record.get("usage_count") or 0)
            total_usage += usage_count
            matches = []
            low_similarity_used = False
            for match in record.get("matches", []):
                flow_uuid = match.get("uuid") or match.get("flow_uuid")
                combined_name = _compose_candidate_name(match)
                if match.get("reasoning") == "low_similarity_included":
                    low_similarity_used = True
                LOGGER.debug(
                    "process_fetcher.dispatch",
                    flow_uuid=flow_uuid,
                    enabled=process_fetch_cfg.enabled,
                )
                process_refs, reference_reason = fetcher.fetch(flow_uuid, combined_name) if flow_uuid else ([], "process_fetch_missing_flow_uuid")
                match_entry = {
                    "flow_uuid": flow_uuid,
                    "flow_name": combined_name,
                    "similarity": match.get("similarity"),
                    "geography": match.get("geography"),
                    "classification": match.get("classification"),
                    "treatment_standards_routes": match.get("treatment_standards_routes"),
                    "mix_and_location_types": match.get("mix_and_location_types"),
                    "flow_properties": match.get("flow_properties"),
                    "version": match.get("version"),
                    "general_comment": match.get("general_comment"),
                    "reasoning": match.get("reasoning"),
                    "process_reference_note": reference_reason,
                    "process_reference_count": len(process_refs),
                    "tiangong_process_references": process_refs,
                }
                matches.append(match_entry)
            notes = list(record.get("notes") or [])
            if not process_fetch_cfg.enabled:
                notes.append("tiangong_process_lookup_missing")
            elif config.process_fetch.enabled and (not config.process_fetch.search_tool_name or not config.process_fetch.crud_tool_name):
                notes.append("process_tool_not_configured")
            if low_similarity_used:
                notes.append("low_similarity")
            unique_reasons = {entry.get("process_reference_note") for entry in matches if entry.get("process_reference_note")}
            for reason in unique_reasons:
                notes.append(reason)
            results.append(
                {
                    "flow_uuid": record.get("flow_uuid"),
                    "flow_name": record.get("flow_name"),
                    "flow_type": record.get("flow_type"),
                    "classification": record.get("classification"),
                    "flow_property": record.get("flow_property"),
                    "synonyms": record.get("synonyms") or [],
                    "usage_count": usage_count,
                    "exchange_occurrences": record.get("exchange_occurrences"),
                    "detail_reference": {
                        "type": "local_jsonl",
                        "path": str(detail_reference_path),
                        "lookup_field": "flow_uuid",
                    },
                    "tiangong_matches": matches,
                    "unmatched_notes": record.get("unmatched", []),
                    "errors": record.get("errors", []),
                    "notes": notes,
                }
            )
            progress_stage.advance()
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": {
            "flow_dir": str(config.flow_dir),
            "process_dir": str(config.process_dir),
            "flow_usage_details": str(config.detail_path),
            "flow_search_results": str(config.search_output_path),
            "flow_types": list(NON_ELEMENTARY_TYPES),
            "git_revision": _detect_git_revision(Path.cwd()),
        },
        "summary": {
            **stats,
            "flow_count": len(results),
            "total_usage_count": total_usage,
            "search_enabled": config.enable_search,
            "process_reference_enabled": config.process_fetch.enabled,
        },
        "results": results,
    }
    config.final_output_path.parent.mkdir(parents=True, exist_ok=True)
    with config.final_output_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
    LOGGER.info(
        "ecoinvent_compare.completed",
        output=str(config.final_output_path),
        flow_count=len(results),
        search_enabled=config.enable_search,
        process_reference_enabled=config.process_fetch.enabled,
    )
    progress_stage.finish()
    return payload


def _read_jsonl(path: Path) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            stripped = line.strip()
            if not stripped:
                continue
            try:
                payload = json.loads(stripped)
            except json.JSONDecodeError as exc:
                LOGGER.warning("jsonl.decode_failed", path=str(path), error=str(exc))
                continue
            if isinstance(payload, dict):
                yield payload
