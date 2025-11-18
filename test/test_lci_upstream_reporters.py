"""Tests for upstream prioritisation reporters."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from tiangong_lca_spec.lci_analysis.upstream.models import (
    ActionItem,
    LifecycleFlowPrioritizationResult,
    PrioritySlice,
    UnknownClassification,
)
from tiangong_lca_spec.lci_analysis.upstream.reporters import write_summary_excel


def test_write_summary_excel_produces_expected_sheets(tmp_path: Path) -> None:
    raw_slice = PrioritySlice(
        dataset_uuid="dataset",
        dataset_name="dataset",
        exchange_name="水泥",
        exchange_name_zh="水泥",
        exchange_name_en="cement",
        flow_name_zh="水泥",
        flow_name_en="cement",
        flow_uuid="flow-uuid",
        flow_role="raw_material",
        unit_family="mass",
        reference_unit="kg",
        total_amount=100.0,
        share=0.8,
        cumulative_share=0.8,
        flow_type="Product flow",
        classification_confidence=0.9,
        rationale="rule",
        exchanges=[],
    )
    downstream_slice = PrioritySlice(
        dataset_uuid="dataset",
        dataset_name="dataset",
        exchange_name="CO2",
        exchange_name_zh="二氧化碳",
        exchange_name_en="CO2",
        flow_name_zh="二氧化碳",
        flow_name_en="CO2",
        flow_uuid="flow-co2",
        flow_role="waste",
        flow_type="Waste flow",
        unit_family="mass",
        reference_unit="kg",
        total_amount=0.2,
        share=0.2,
        cumulative_share=0.2,
        downstream_path="unknown",
        downstream_action="check emission permit",
        exchanges=[],
    )
    unknown_entry = UnknownClassification(exchange_name="detergent", dataset_uuid="dataset", reason="missing rule")
    action = ActionItem(priority="high", type="upstream", summary="trace cement inputs", evidence=["share≈0.8"])
    result = LifecycleFlowPrioritizationResult(
        raw_materials=[raw_slice],
        energy=[],
        auxiliaries=[],
        downstream_outputs=[downstream_slice],
        unknown_classifications=[unknown_entry],
        actions=[action],
        notes=["unit_family_missing"],
        metadata={"run_id": "demo", "schema_version": 2},
    )

    output_path = write_summary_excel(result, tmp_path)

    assert output_path.exists()
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Upstream", "Downstream", "Unknown", "Actions", "Notes", "Metadata"]
    upstream = workbook["Upstream"]
    assert upstream["A1"].value == "dataset_uuid"
    assert upstream["C2"].value == "mass"
    assert upstream["M2"].value.endswith("%")
    assert upstream["O1"].value == "reference_process_count"
    assert upstream["O2"].value is None

    downstream = workbook["Downstream"]
    assert downstream["A1"].value == "dataset_uuid"
    assert downstream["A2"].value == "dataset"
    assert downstream["P1"].value == "downstream_path"
    assert downstream["P2"].value == "unknown"
    assert downstream["Q2"].value == "check emission permit"

    unknown_ws = workbook["Unknown"]
    assert unknown_ws.max_row == 2

    actions_ws = workbook["Actions"]
    assert "trace cement inputs" in actions_ws["C2"].value

    notes_ws = workbook["Notes"]
    assert notes_ws["A2"].value == "unit_family_missing"

    metadata_ws = workbook["Metadata"]
    assert metadata_ws.max_row == 3
